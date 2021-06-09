import openpyxl

excel_file = openpyxl.load_workbook("inventory.xlsx")
product_list = excel_file["Sheet1"]

dict1 = {}
dict2 = {}
dict3 = {}

for row in range(2, product_list.max_row + 1):
    product_num = product_list.cell(row, 1).value
    inventory = product_list.cell(row, 2).value
    price = product_list.cell(row, 3).value
    supplier_name = product_list.cell(row, 4).value
    inventory_price = product_list.cell(row, 5)

    #EX1: list each  supplier with product number
    if supplier_name in dict1:
        current_pd_nb = dict1.get(supplier_name)
        dict1[supplier_name] = current_pd_nb + 1
    else:
        dict1[supplier_name] = 1

    #EX2: list products with inventory less than 10
    if inventory < 10:
        dict2[int(product_num)] = int(inventory)

    #EX3: list each supplier with total inventory
    if supplier_name in dict3:
        current_total_inv = dict3.get(supplier_name)
        dict3[supplier_name] = current_total_inv + (price*inventory)
    else:
        dict3[supplier_name] = price * inventory

    # Ex4: Add new column: inventory value for each product
    inventory_price.value = price * inventory


excel_file.save("spreadshett_tot_val.xlsx")


print(dict1)
print(dict2)
print(dict3)



